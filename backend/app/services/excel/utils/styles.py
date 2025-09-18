from __future__ import annotations

import re
from collections import Counter
from typing import Optional, Dict, Any, Tuple

import pandas as pd
from openpyxl import load_workbook

# ---------------- Styles & header inference ---------------- #

def hex_from_fill(fill) -> Optional[str]:
    """
    Return ARGB hex like 'FF00FF00' (alpha + rgb) or None if no solid fill.
    openpyxl stores theme/tint sometimes; this covers the common case.
    """
    try:
        if fill and getattr(fill, "patternType", None) == "solid":
            fg = getattr(fill, "fgColor", None)
            if fg and getattr(fg, "rgb", None):
                return fg.rgb.upper()
    except Exception:
        pass
    return None


def _row_features(ws, r: int, max_cols: int = 100) -> Dict[str, Any]:
    """
    Build simple features to score whether row r looks like a header row.
    """
    values = []
    bolds = []
    fills = []
    for c in range(1, min(ws.max_column, max_cols) + 1):
        cell = ws.cell(row=r, column=c)
        values.append(cell.value)
        bolds.append(bool(getattr(getattr(cell, "font", None), "bold", False)))
        fills.append(hex_from_fill(getattr(cell, "fill", None)))

    vals = [v for v in values if v is not None and str(v).strip() != ""]
    svals = [str(v).strip() for v in vals]
    has_alpha = sum(bool(re.search(r"[A-Za-z]", v)) for v in svals)

    return {
        "row": r,
        "n_nonempty": len(vals),
        "share_nonempty": len(vals) / max(1, min(ws.max_column, max_cols)),
        "share_alpha": has_alpha / max(1, len(svals)),  # how text-like
        "n_unique": len(set(svals)),
        "share_bold": sum(bolds) / max(1, len(bolds)),
        "fill_mode": Counter([f for f in fills if f]).most_common(1)[0][0] if any(fills) else None,
        "values_preview": svals[:10],
    }


def infer_header_row(ws, search_top_n: int = 20) -> int:
    """
    Score the first `search_top_n` rows and pick the one that most looks like a header.
    Heuristics:
      - many non-empty cells
      - mostly alphabetic text
      - high uniqueness (column names tend to be unique)
      - some bold or colored cells help
    """
    candidates = []
    for r in range(1, min(ws.max_row, search_top_n) + 1):
        f = _row_features(ws, r)
        score = (
            2.0 * f["share_nonempty"] +
            1.0 * f["share_alpha"] +
            0.5 * (f["n_unique"] / max(1, f["n_nonempty"])) +
            0.3 * f["share_bold"]
        )
        if f["fill_mode"]:
            score += 0.2
        candidates.append((score, f))

    best = max(candidates, key=lambda x: x[0])[1]
    return best["row"]


def read_excel_with_styles(path: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[Tuple[int, int], str]]:
    """
    Returns: (df, color_map) where
      - df is a pandas DataFrame with inferred header
      - color_map is a dict[(row_idx, col_idx)] = hex ARGB like 'FFFF0000'
        (0-based indices aligned to df, not Excel)
    Note: works best for .xlsx
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = infer_header_row(ws)
    df = pd.read_excel(path, sheet_name=ws.title, header=header_row - 1, engine="openpyxl")

    # Build a color map for data rows beneath header
    color_map: Dict[Tuple[int, int], str] = {}
    start_excel_row = header_row + 1
    for r in range(start_excel_row, header_row + 1 + len(df)):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            hexargb = hex_from_fill(cell.fill)
            if hexargb:
                # map to df coords (0-based)
                df_r = (r - header_row - 1)
                df_c = (c - 1)
                color_map[(df_r, df_c)] = hexargb

    return df, color_map
