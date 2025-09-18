import re
from collections import Counter
from typing import Optional, Tuple, Dict, Any, List

import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import process, fuzz  # for fuzzy column-name matching (pip install rapidfuzz)

# --------- Utilities

def _hex_from_fill(fill) -> Optional[str]:
    """
    Return ARGB hex like 'FF00FF00' (alpha + rgb) or None if no solid fill.
    openpyxl stores theme/tint sometimes; this covers the common case.
    """
    try:
        if fill and fill.patternType == "solid" and fill.fgColor and fill.fgColor.rgb:
            return fill.fgColor.rgb.upper()
    except Exception:
        pass
    return None

def _row_features(ws, r, max_cols=100) -> Dict[str, Any]:
    """
    Build simple features to score whether row r looks like a header row.
    """
    values = []
    types = []
    bolds = []
    fills = []
    for c in range(1, min(ws.max_column, max_cols) + 1):
        cell = ws.cell(row=r, column=c)
        values.append(cell.value)
        types.append(type(cell.value).__name__)
        bolds.append(bool(getattr(getattr(cell, "font", None), "bold", False)))
        fills.append(_hex_from_fill(getattr(cell, "fill", None)))

    vals = [v for v in values if v is not None and str(v).strip() != ""]
    svals = [str(v).strip() for v in vals]
    has_commons = sum(bool(re.search(r"[A-Za-z]", v)) for v in svals)

    return {
        "row": r,
        "n_nonempty": len(vals),
        "share_nonempty": len(vals) / max(1, min(ws.max_column, max_cols)),
        "share_alpha": has_commons / max(1, len(svals)),      # how text-like
        "n_unique": len(set(svals)),
        "share_bold": sum(bolds) / max(1, len(bolds)),
        "fill_mode": Counter([f for f in fills if f]).most_common(1)[0][0] if any(fills) else None,
        "values_preview": svals[:10],
    }

def infer_header_row(ws, search_top_n=20) -> int:
    """
    Score the first `search_top_n` rows and pick the one that most looks like a header.
    Heuristics:
      - many non-empty cells
      - mostly alphabetic text
      - high uniqueness (column names tend to be unique)
      - some bold or colored cells help
    """
    candidates = []
    for r in range(1, min(ws.max_row, search_top_n) + 1):
        f = _row_features(ws, r)
        score = (
            2.0 * f["share_nonempty"] +
            1.0 * f["share_alpha"] +
            0.5 * (f["n_unique"] / max(1, f["n_nonempty"])) +
            0.3 * f["share_bold"]
        )
        # small bonus if the row has a uniform fill (common in headers)
        if f["fill_mode"]:
            score += 0.2
        candidates.append((score, f))

    best = max(candidates, key=lambda x: x[0])[1]
    return best["row"]

def read_excel_with_styles(path: str, sheet_name: Optional[str] = None):
    """
    Returns: (df, color_map) where
      - df is a pandas DataFrame with inferred header
      - color_map is a dict[(row_idx, col_idx)] = hex ARGB like 'FFFF0000'
        (0-based indices aligned to df, not Excel)
    Note: works best for .xlsx
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = infer_header_row(ws)
    # read sheet with pandas using header at (header_row-1) zero-based
    df = pd.read_excel(path, sheet_name=ws.title, header=header_row-1, engine="openpyxl")

    # Build a color map for data rows beneath header
    color_map = {}
    start_excel_row = header_row + 1
    for r in range(start_excel_row, header_row + 1 + len(df)):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            hexargb = _hex_from_fill(cell.fill)
            if hexargb:
                # map to df coords (0-based)
                df_r = (r - header_row - 1)
                df_c = (c - 1)
                color_map[(df_r, df_c)] = hexargb

    return df, color_map

def pick_island_column(df: pd.DataFrame, candidates: List[str] = None) -> Optional[str]:
    """
    Try to find an 'Island' / region column even if named weirdly.
    """
    if candidates is None:
        candidates = [
            "island", "islands", "region", "county", "area", "province",
            "location", "territory", "state"
        ]
    # direct match first
    for col in df.columns:
        if str(col).strip().lower() in candidates:
            return col

    # fuzzy match
    choices = list(map(str, df.columns))
    match, score, _ = process.extractOne(
        "island",
        choices,
        scorer=fuzz.WRatio
    )
    return match if score >= 80 else None

# --------- Example “analysis” entry point

def analyze_excel(path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    df, color_map = read_excel_with_styles(path, sheet_name=sheet_name)

    # Find an island-ish column and run a quick groupby
    island_col = pick_island_column(df)
    by_island = None
    if island_col:
        # example: count rows per island and any numeric summary
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if numeric_cols:
            by_island = df.groupby(island_col)[numeric_cols].agg(["count", "mean", "sum"])
        else:
            by_island = df.groupby(island_col).size().to_frame("rows")

    # Example: count background colors used in the data region
    color_counts = Counter(color_map.values())
    most_common_colors = color_counts.most_common(10)

    return {
        "columns": list(df.columns),
        "preview": df.head(10).to_dict(orient="records"),
        "island_column": island_col,
        "by_island_summary": None if by_island is None else by_island.head(20).to_dict(),
        "color_usage_top10": most_common_colors,
        "notes": [
            "Header row inferred with heuristics (non-empty density, text-likeness, bold, fill).",
            "Colors are ARGB hex from Excel fills; theme/tint nuances may need extra handling.",
            "Rename/clean headers with libraries like `pyjanitor` or `python-slugify` if needed."
        ]
    }
